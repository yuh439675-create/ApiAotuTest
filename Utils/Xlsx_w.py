from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os
from Config.config import Config
from datetime import datetime
import re

# 配置参数
FILE_PATH = os.path.join(Config.Xlsx_Words, "国外短剧自动化问题_自动编号版.xlsx")
HEADER = ["样式列", "编号", "语言", "剧名", "整剧价格", "每集价格", "大于100MB的问题剧集", "是否已重新上传", "是否收费",
          "是否上架", "是否完结", "大小", "上传时间"]
COLUMN_WIDTHS = [1, 8, 10, 30, 10, 10, 35, 24, 10, 10, 10, 8, 20]

# 样式定义
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
WHITE_BOLD_FONT = Font(color="FFFFFF", bold=True)
LIGHT_BLUE_FILL = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
LIGHT_ORANGE_FILL = PatternFill(start_color="FFF0E0", end_color="FFF0E0", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def init_or_load_workbook():
    if os.path.exists(FILE_PATH):
        wb = load_workbook(FILE_PATH)
        ws = wb.active
        # 清理现有的空行，只保留有数据的行
        _clean_existing_data(ws)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _initialize_worksheet(ws)
    return wb, ws


def _clean_existing_data(ws):
    """清理现有数据，移除空行，只保留有实际数据的行"""
    rows_to_delete = []

    # 从第2行开始检查，第1行是表头
    for row in range(2, ws.max_row + 1):
        # 检查B列（编号）是否有数据，如果有数据则保留
        id_cell = ws[f"B{row}"]
        if not id_cell.value or id_cell.value == "":
            # 如果B列没有数据，检查是否是日期分隔行
            a_cell = ws[f"A{row}"]
            if a_cell.value and "数据结束" in str(a_cell.value):
                # 保留日期分隔行
                continue
            else:
                # 标记为要删除的空行
                rows_to_delete.append(row)

    # 从后往前删除行，避免索引变化
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)

    # 重新设置冻结窗格
    ws.freeze_panes = "A2"


def _initialize_worksheet(ws):
    """初始化工作表表头和样式"""
    ws.freeze_panes = "A2"

    # 设置表头行高
    ws.row_dimensions[1].height = 20  # 表头行高

    # 设置表头
    for col_idx, header in enumerate(HEADER, 1):
        col_letter = get_column_letter(col_idx)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = WHITE_BOLD_FONT
        cell.fill = GREEN_FILL
        cell.alignment = CENTER_ALIGN
        ws.column_dimensions[col_letter].width = COLUMN_WIDTHS[col_idx - 1]

    # 样式列合并 - 只在创建新文件时合并
    if not os.path.exists(FILE_PATH):
        ws.merge_cells('A1:A501')

    # 设置数据行行高（第2行到第501行）
    for row in range(2, 502):
        ws.row_dimensions[row].height = 20  # 数据行高

    # 设置第一列的自定义背景色（除了表头）
    CUSTOM_FIRST_COLUMN_FILL = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # 浅紫色
    for row in range(2, 502):
        cell = ws[f"A{row}"]
        cell.fill = CUSTOM_FIRST_COLUMN_FILL
        cell.alignment = CENTER_ALIGN

    # 设置H列的浅红色背景
    for row in range(2, 502):
        cell = ws[f"H{row}"]
        cell.fill = LIGHT_RED_FILL


def get_last_data_date(ws):
    """获取最后一条数据的日期"""
    last_row = ws.max_row
    for row in range(last_row, 1, -1):
        upload_time_cell = ws[f"L{row}"]
        if upload_time_cell.value and isinstance(upload_time_cell.value, str):
            try:
                date_str = upload_time_cell.value.split(" ")[0]
                return datetime.strptime(date_str, "%Y/%m/%d").date()
            except:
                continue
    return None


def get_last_id(ws):
    """获取最后一个编号"""
    last_row = ws.max_row
    for row in range(last_row, 1, -1):
        id_cell = ws[f"B{row}"]
        if id_cell.value and isinstance(id_cell.value, str):
            match = re.match(r"TV_(\d+)", id_cell.value)
            if match:
                return int(match.group(1))
    return 0


def find_next_empty_row(ws):
    """找到下一个空行（从第2行开始找）"""
    for row in range(2, ws.max_row + 2):
        # 检查B列是否有数据
        id_cell = ws[f"B{row}"]
        if not id_cell.value or id_cell.value == "":
            # 检查A列是否是日期分隔行
            a_cell = ws[f"A{row}"]
            if a_cell.value and "数据结束" in str(a_cell.value):
                continue  # 跳过日期分隔行
            return row
    return ws.max_row + 1


def insert_date_separator(ws, last_data_date, current_date):
    """插入日期分隔行"""
    next_row = find_next_empty_row(ws)

    # 安全地取消合并单元格
    merged_ranges_to_remove = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row <= next_row <= merged_range.max_row:
            merged_ranges_to_remove.append(merged_range)

    for merged_range in merged_ranges_to_remove:
        try:
            # 使用更安全的方法取消合并
            ws.unmerge_cells(str(merged_range))
        except KeyError as e:
            print(f"警告：取消合并单元格时遇到不存在的单元格 {e}，继续执行")
            # 强制从合并单元格列表中移除
            if merged_range in ws.merged_cells:
                ws.merged_cells.remove(merged_range)

    # 合并单元格
    ws.merge_cells(f"A{next_row}:L{next_row}")

    # 获取合并单元格的锚点单元格（左上角单元格）
    separator_cell = ws[f"A{next_row}"]
    separator_cell.value = f"--- {last_data_date.strftime('%Y/%m/%d')} 数据结束 ---"
    separator_cell.fill = LIGHT_ORANGE_FILL
    separator_cell.alignment = CENTER_ALIGN

    return next_row + 1


def append_data(ws, data_list, current_datetime=None):
    num_data = 10
    """追加数据到工作表"""
    if len(data_list) != num_data:
        raise ValueError(f"数据列表必须包含{num_data}个元素（语言-是否完结），当前仅{len(data_list)}个")

    if not current_datetime:
        current_datetime = datetime.now().strftime("%Y/%m/%d %H:%M")

    current_date = datetime.strptime(current_datetime.split(" ")[0], "%Y/%m/%d").date()
    last_data_date = get_last_data_date(ws)

    # 如果需要插入日期分隔行
    insert_row = find_next_empty_row(ws)
    if last_data_date and current_date > last_data_date:
        insert_row = insert_date_separator(ws, last_data_date, current_date)

    # 生成新编号
    last_id = get_last_id(ws)
    new_id = f"TV_{last_id + 1:03d}"

    # 构建完整的数据行
    full_data = [new_id] + data_list + [current_datetime]

    # 批量设置数据和样式
    for col_idx, value in enumerate(full_data, 2):
        col_letter = get_column_letter(col_idx)
        cell = ws[f"{col_letter}{insert_row}"]
        cell.value = value
        cell.alignment = CENTER_ALIGN
        # 非H列应用浅蓝色，H列保持浅红色
        if col_idx != 8:
            cell.fill = LIGHT_BLUE_FILL

    print(f"已插入数据：行{insert_row}, 编号{new_id}")


def append_multiple_data(ws, data_batch, datetime_batch=None):
    """
    批量追加数据到工作表，提高大量数据写入性能
    """
    if not data_batch:
        return
        
    if datetime_batch is None:
        datetime_batch = [datetime.now().strftime("%Y/%m/%d %H:%M")] * len(data_batch)
    elif len(datetime_batch) != len(data_batch):
        raise ValueError("数据批次和时间批次长度不匹配")
        
    # 批量处理数据，减少对Excel文件的访问次数
    for i, data_list in enumerate(data_batch):
        append_data(ws, data_list, datetime_batch[i])


if __name__ == '__main__':
    wb, ws = init_or_load_workbook()

    new_data4 = ["葡萄牙语", "CEO and I A Bittersweet Goodbye", "100$", "4.5$", "CEO and I A Bittersweet Goodbye 1", "",
                 "是", "是", "是", "12MB"]
    append_data(ws, new_data4, "2025/11/7 11:50")

    ew_data4 = ["葡萄牙语", "CEO and I A Bittersweet Goodbye", "100$", "4.5$", "CEO and I A Bittersweet Goodbye 1", "",
                "是", "是", "是", "10MB"]
    append_data(ws, new_data4, "2025/11/8 11:50")

    wb.save(FILE_PATH)
    print(f"表格已更新：{FILE_PATH}")