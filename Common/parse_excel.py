from openpyxl import load_workbook


class ParseExcel(object):
    def __init__(self, excelPath, sheetName):
        self.wb = load_workbook(excelPath)
        self.sheet = self.wb[sheetName]
        self.maxRowNum = self.sheet.max_row
        self.maxColNum = self.sheet.max_column  # 新增：获取最大列数

    def getDataFromSheet(self):
        # 这个方法保持不变，用于读取指定列数的数据
        dataList = []
        print(self.sheet.rows)
        for line in self.sheet.rows:
            tmplist = [line[i].value for i in range(self.sheet.max_column)]  # 使用列表推导式简化代码
            dataList.append(tmplist)
        return dataList[1:]  # 假设您要跳过前三行（标题等）

    def getAllDataFromSheet(self):
        # 这个新方法用于读取所有列的数据
        global tmplist
        dataList = []
        for line in self.sheet.rows:
            tmplist = [cell.value for cell in line]  # 读取整行的所有单元格
            dataList.append(tmplist)
        data = dataList[1:len(tmplist)]
        # 使用新方法来读取所有列的数据
        list_data = []
        for row in data:
            list_data.append(
                {
                    '编号': row[0],
                    'BUG描述': row[1],
                    '发现日期': row[2],
                    '发现人': row[3],
                    '状态': row[4],
                    '优先级(致命/严重/一般/轻微)': row[5],
                    '所属模块': row[6],
                    '处理人': row[7],
                    '处理日期': row[8],
                    '是/否(已解决)': row[9],
                    '备注': row[10],
                }
            )
        l = [(case['编号'], case['BUG描述'], case['发现日期'], case['发现人'], case['状态'],
              case['优先级(致命/严重/一般/轻微)'],
              case['所属模块'], case['处理人'], case['处理日期'], case['是/否(已解决)'],
              case['备注']) for case in list_data]
        return l


# __main__ 块保持不变
if __name__ == '__main__':
    excel_path = r'C:\Users\28242\Desktop\PWA-产出物\BUG记录\总的\总体-PWA-BUG管理.xlsx'
    sheet_name = 'Sheet1'
    parser = ParseExcel(excel_path, sheet_name)
    data1 = parser.getAllDataFromSheet()
    data2 = parser.getDataFromSheet()
    print(data1)

# from datetime import datetime
# from typing import List, Dict, Tuple, Any
# from openpyxl import load_workbook
# from openpyxl.worksheet.worksheet import Worksheet
#
#
# class ExcelParser:
#     """Excel 文件解析器，支持动态列映射和类型转换"""
#
#     # 列定义模板（可根据实际Excel表头修改）
#     COLUMN_MAPPING = {
#         '编号': {'index': 0, 'type': str},
#         'BUG描述': {'index': 1, 'type': str},
#         '发现日期': {'index': 2, 'type': str},
#         '发现人': {'index': 3, 'type': str},
#         '状态': {'index': 4, 'type': str},
#         '优先级(致命/严重/一般/轻微)': {'index': 5, 'type': str},
#         '所属模块': {'index': 6, 'type': str},
#         '处理人': {'index': 7, 'type': str},
#         '处理日期': {'index': 8, 'type': str},
#         '是/否(已解决)': {'index': 9, 'type': str},
#         '备注': {'index': 10, 'type': str},
#     }
#
#     def __init__(self, file_path: str, sheet_name: str = 'Sheet1'):
#         """
#         初始化解析器
#         :param file_path: Excel文件路径
#         :param sheet_name: 工作表名称，默认为Sheet1
#         """
#         self.wb = load_workbook(file_path, read_only=True, data_only=True)
#         self.sheet = self.wb[sheet_name]
#         self.max_row = self.sheet.max_row
#         self._validate_columns()
#
#     def _validate_columns(self):
#         """验证工作表列是否匹配预设格式"""
#         header = next(self.sheet.iter_rows(min_row=1, max_row=1, values_only=True))
#         for col_name, config in self.COLUMN_MAPPING.items():
#             if header[config['index']] != col_name:
#                 raise ValueError(f"列[{config['index']}]应为'{col_name}'，实际为'{header[config['index']]}'")
#
#     def _convert_type(self, value: Any, col_type: type) -> Any:
#         """类型转换处理器"""
#         try:
#             if col_type == datetime and isinstance(value, str):
#                 return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
#             if value is None and col_type != str:
#                 return None
#             return col_type(value)
#         except (ValueError, TypeError) as e:
#             print(f"类型转换失败: {value} -> {col_type.__name__}, 错误: {str(e)}")
#             return value
#
#     def parse_data(self, skip_rows: int = 1) -> List[Dict]:
#         """
#         解析完整数据
#         :param skip_rows: 跳过的标题行数，默认为1
#         :return: 字典列表，每个字典代表一行数据
#         """
#         result = []
#         for row in self.sheet.iter_rows(min_row=skip_rows + 1, values_only=True):
#             try:
#                 item = {
#                     col_name: self._convert_type(row[config['index']], config['type'])
#                     for col_name, config in self.COLUMN_MAPPING.items()
#                 }
#                 result.append(item)
#             except IndexError as e:
#                 print(f"数据行缺失列: {str(e)}，行内容: {row}")
#         return result
#
#     def get_tuples(self, skip_rows: int = 1) -> List[Tuple]:
#         """获取元组形式数据，适用于数据库批量插入"""
#         return [
#             tuple(item.values())
#             for item in self.parse_data(skip_rows)
#         ]
#
#     def __enter__(self):
#         return self
#
#     def __exit__(self, exc_type, exc_val, exc_tb):
#         self.wb.close()
#
#
# if __name__ == '__main__':
#     # 使用示例
#     excel_path = r'C:\Users\28242\Desktop\PWA-产出物\BUG记录\总的\总体-PWA-BUG管理.xlsx'
#
#     try:
#         with ExcelParser(excel_path) as parser:
#
#             # 获取字典格式数据
#             # dict_data = parser.parse_data()
#             # for ros in range(parser.max_row-1):
#             #     print(f"字典格式第{ros+1}条数据:", dict_data[ros])
#
#             # # 获取元组格式数据
#             tuple_data = parser.get_tuples()
#             for ros in range(parser.max_row-1):
#                 print(f"\n元组格式第{ros+1}条数据:", tuple_data[ros])
#
#     except FileNotFoundError:
#         print(f"文件未找到: {excel_path}")
#     except KeyError as e:
#         print(f"工作表不存在: {str(e)}")
#     except ValueError as e:
#         print(f"列验证失败: {str(e)}")
